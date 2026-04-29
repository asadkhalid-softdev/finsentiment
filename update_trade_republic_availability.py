import argparse
import re
import shutil
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from difflib import SequenceMatcher
from urllib.parse import quote

from openpyxl import load_workbook
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
from playwright.sync_api import sync_playwright


DEFAULT_FILE = "stocks_metrics.xlsx"
DEFAULT_SHEET = "Metrics"
DEFAULT_PROFILE_DIR = r"C:\Users\asad_\.playwright-mcp\profile"
TR_STOCK_SEARCH_URL = "https://app.traderepublic.com/browse/stock?q={query}"


LEGAL_SUFFIX_RE = re.compile(
    r"\b("
    r"incorporated|inc|corporation|corp|company|co|limited|ltd|holdings?|group|"
    r"plc|public limited company|aktiengesellschaft|ag|se|s\.a\.|sa|a/s|"
    r"n\.v\.|nv|p\.l\.c\.|tbk|kgaa|gmbh|asa|oyj|a/s"
    r")\b\.?",
    re.IGNORECASE,
)
ISIN_RE = re.compile(r"^[A-Z]{2}[A-Z0-9]{9}[0-9]$")
RESULT_COUNT_RE = re.compile(r"(\d+)\s+results?\s+found", re.IGNORECASE)


@dataclass
class SearchResult:
    name: str
    isin: str
    confidence: int


def normalize_name(value: str) -> str:
    text = str(value or "").strip()
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"\s+R$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\bRegistered Shares.*$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\bReg\.?\s+Shares.*$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\bRegistered\s+Shs.*$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\bReg\.?\s+Shs.*$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\bShares\s+(Cl\.?|Class|DL|o\.N\.).*$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\([^)]*\)", " ", text)
    text = text.replace("&", " and ")
    text = LEGAL_SUFFIX_RE.sub(" ", text)
    text = re.sub(r"[^a-zA-Z0-9]+", " ", text).lower()
    text = re.sub(r"^the\s+", "", text)
    return re.sub(r"\s+", " ", text).strip()


def search_queries(company_name: str, ticker: str = "") -> list[str]:
    raw = re.sub(r"\s+", " ", str(company_name or "")).strip()
    no_parens = re.sub(r"\s*\([^)]*\)\s*", " ", raw).strip()
    display_clean = LEGAL_SUFFIX_RE.sub(" ", no_parens.replace(",", " "))
    display_clean = re.sub(r"\s+", " ", display_clean).strip()
    normalized = normalize_name(raw)
    ticker_base = str(ticker or "").strip().split(".")[0]
    queries = [display_clean, normalized, no_parens, raw, ticker_base]

    seen = set()
    unique_queries = []
    for query in queries:
        query = re.sub(r"\s+", " ", query).strip()
        if query and query.lower() not in seen:
            seen.add(query.lower())
            unique_queries.append(query)
    return unique_queries


def acronym(value: str) -> str:
    words = normalize_name(value).split()
    ignored = {"and", "of", "the"}
    return "".join(word[0] for word in words if word not in ignored and word)


def confidence_score(expected: str, result_name: str, ticker: str = "") -> int:
    left = normalize_name(expected)
    right = normalize_name(result_name)
    if not left or not right:
        return 0
    ticker_base = normalize_name(str(ticker or "").split(".")[0])
    if ticker_base and right == ticker_base:
        return 100
    if acronym(expected) and right == acronym(expected):
        return 100
    if left == right:
        return 100

    ratio = SequenceMatcher(None, left, right).ratio()
    left_tokens = set(left.split())
    right_tokens = set(right.split())
    if left_tokens and left_tokens <= right_tokens:
        return 100
    if right_tokens and right_tokens <= left_tokens and len(right_tokens) >= 2:
        return 100
    overlap = len(left_tokens & right_tokens) / max(1, len(left_tokens | right_tokens))
    subset_bonus = 0.12 if left_tokens <= right_tokens or right_tokens <= left_tokens else 0
    return min(100, round((ratio * 0.7 + overlap * 0.3 + subset_bonus) * 100))


def parse_stock_results(body_text: str, expected_name: str, ticker: str = "") -> list[SearchResult]:
    lines = [line.strip() for line in body_text.splitlines() if line.strip()]
    results = []

    for index, line in enumerate(lines):
        if ISIN_RE.match(line) and index > 0:
            name = lines[index - 1]
            if name.lower() in {"security", "stocks", "crypto", "etfs", "bonds", "derivatives"}:
                continue
            results.append(
                SearchResult(
                    name=name,
                    isin=line,
                    confidence=confidence_score(expected_name, name, ticker),
                )
            )

    deduped = {}
    for result in results:
        key = (result.name, result.isin)
        if key not in deduped or result.confidence > deduped[key].confidence:
            deduped[key] = result
    return sorted(deduped.values(), key=lambda item: item.confidence, reverse=True)


def find_header_indexes(sheet) -> dict[str, int]:
    headers = {}
    for cell in sheet[1]:
        if cell.value:
            headers[str(cell.value).strip()] = cell.column
    return headers


def ensure_column(sheet, headers: dict[str, int], name: str) -> int:
    if name in headers:
        return headers[name]
    column = sheet.max_column + 1
    sheet.cell(row=1, column=column).value = name
    headers[name] = column
    return column


def wait_for_manual_login(page) -> None:
    print("A browser window is open. Log in to Trade Republic there if needed.")
    print("After the portfolio or app page is visible, press Enter here to continue.")
    page.goto("https://app.traderepublic.com/portfolio", wait_until="domcontentloaded")
    input()


def lookup_company(
    page,
    company_name: str,
    ticker: str,
    threshold: int,
    delay: float,
    settle_ms: int,
    debug_dir: Path | None = None,
    row_number: int | None = None,
) -> tuple[bool, SearchResult | None, str, int]:
    best_result = None
    best_query = ""
    best_count = 0

    for query in search_queries(company_name, ticker):
        page.goto(TR_STOCK_SEARCH_URL.format(query=quote(query, safe="")), wait_until="domcontentloaded")
        if settle_ms:
            page.wait_for_timeout(settle_ms)
        try:
            page.wait_for_function(
                r"""() =>
                    (
                        document.body.innerText.includes("All Products for") &&
                        /\d+\s+results?\s+found/i.test(document.body.innerText)
                    ) ||
                    document.body.innerText.includes("We couldn't find")
                """,
                timeout=15000,
            )
        except PlaywrightTimeoutError:
            pass

        body = page.locator("body").inner_text(timeout=5000)
        count_match = RESULT_COUNT_RE.search(body)
        result_count = int(count_match.group(1)) if count_match else 0
        if result_count > best_count:
            best_count = result_count
            if not best_query:
                best_query = query
        results = parse_stock_results(body, company_name, ticker)

        if results and (best_result is None or results[0].confidence > best_result.confidence):
            best_result = results[0]
            best_query = query
            best_count = result_count

        if debug_dir and result_count == 0 and not results:
            debug_dir.mkdir(parents=True, exist_ok=True)
            safe_query = re.sub(r"[^a-zA-Z0-9._-]+", "_", query)[:80]
            debug_name = f"row_{row_number or 'unknown'}_{safe_query}.txt"
            (debug_dir / debug_name).write_text(
                f"URL: {page.url}\nQUERY: {query}\n\n{body[:8000]}",
                encoding="utf-8",
            )

        if best_result and best_result.confidence >= threshold:
            return True, best_result, best_query, best_count

        if delay:
            time.sleep(delay)

    return False, best_result, best_query, best_count


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Update stocks_metrics.xlsx on_TradeRepublic by checking Trade Republic stock search pages."
    )
    parser.add_argument("--file", default=DEFAULT_FILE)
    parser.add_argument("--sheet", default=DEFAULT_SHEET)
    parser.add_argument("--profile-dir", default=DEFAULT_PROFILE_DIR)
    parser.add_argument("--threshold", type=int, default=82)
    parser.add_argument("--delay", type=float, default=0.25)
    parser.add_argument("--settle-ms", type=int, default=1500, help="Extra wait after each Trade Republic page load.")
    parser.add_argument("--debug-dir", default="", help="Optional directory for page text dumps when no results are parsed.")
    parser.add_argument("--headless", action="store_true")
    parser.add_argument("--login", action="store_true", help="Open the persistent browser profile and wait for manual login.")
    parser.add_argument("--limit", type=int, default=0, help="Optional row limit for a test run.")
    parser.add_argument("--no-backup", action="store_true")
    parser.add_argument("--allow-zero-results", action="store_true", help="Disable the fail-fast guard for empty result pages.")
    args = parser.parse_args()

    workbook_path = Path(args.file)
    if not workbook_path.exists():
        print(f"Workbook not found: {workbook_path}", file=sys.stderr)
        return 1

    if not args.no_backup:
        timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        backup_path = workbook_path.with_name(f"{workbook_path.stem}.backup-before-tr-{timestamp}{workbook_path.suffix}")
        shutil.copy2(workbook_path, backup_path)
        print(f"Backup written: {backup_path}")

    workbook = load_workbook(workbook_path)
    if args.sheet not in workbook.sheetnames:
        print(f"Sheet not found: {args.sheet}", file=sys.stderr)
        return 1
    sheet = workbook[args.sheet]

    headers = find_header_indexes(sheet)
    if "Company Name" not in headers:
        print("Missing required column: Company Name", file=sys.stderr)
        return 1

    company_col = headers["Company Name"]
    ticker_col = headers.get("Ticker")
    on_tr_col = ensure_column(sheet, headers, "on_TradeRepublic")
    matched_name_col = ensure_column(sheet, headers, "tr_matched_name")
    confidence_col = ensure_column(sheet, headers, "tr_match_confidence")
    query_col = ensure_column(sheet, headers, "tr_matched_query")
    isin_col = ensure_column(sheet, headers, "tr_matched_isin")

    data_rows = []
    for row_number in range(2, sheet.max_row + 1):
        company_name = sheet.cell(row=row_number, column=company_col).value
        if company_name:
            ticker = sheet.cell(row=row_number, column=ticker_col).value if ticker_col else ""
            data_rows.append((row_number, str(company_name).strip(), str(ticker or "").strip()))
    if args.limit:
        data_rows = data_rows[: args.limit]

    profile_dir = Path(args.profile_dir)
    with sync_playwright() as playwright:
        context = playwright.chromium.launch_persistent_context(
            str(profile_dir),
            channel="chrome",
            headless=args.headless,
            viewport={"width": 1400, "height": 950},
        )
        page = context.pages[0] if context.pages else context.new_page()

        if args.login:
            wait_for_manual_login(page)

        true_count = 0
        false_count = 0
        zero_result_rows = 0
        abort_due_to_empty_pages = False
        debug_dir = Path(args.debug_dir) if args.debug_dir else None
        for index, (row_number, company_name, ticker) in enumerate(data_rows, start=1):
            try:
                is_available, result, query, result_count = lookup_company(
                    page,
                    company_name,
                    ticker,
                    args.threshold,
                    args.delay,
                    args.settle_ms,
                    debug_dir,
                    row_number,
                )
            except Exception as exc:
                print(f"[{index}/{len(data_rows)}] row {row_number}: error for {company_name}: {exc}")
                is_available, result, query, result_count = False, None, "", 0

            sheet.cell(row=row_number, column=on_tr_col).value = bool(is_available)
            sheet.cell(row=row_number, column=matched_name_col).value = result.name if result else ""
            sheet.cell(row=row_number, column=confidence_col).value = result.confidence if result else 0
            sheet.cell(row=row_number, column=query_col).value = query
            sheet.cell(row=row_number, column=isin_col).value = result.isin if result else ""

            true_count += int(is_available)
            false_count += int(not is_available)
            zero_result_rows += int(result is None and result_count == 0)
            matched = result.name if result else "no parsed result"
            score = result.confidence if result else 0
            print(
                f"[{index}/{len(data_rows)}] row {row_number}: "
                f"{company_name} -> {is_available} "
                f"({matched}, score={score}, results={result_count}, query={query or 'n/a'})"
            )

            guard_window = min(5, len(data_rows))
            if not args.allow_zero_results and index == guard_window and zero_result_rows == guard_window:
                abort_due_to_empty_pages = True
                print(
                    "\nAborting without saving: the first rows all returned empty/unparsed Trade Republic pages. "
                    "This usually means the browser profile is not logged in, the page did not load, "
                    "or Trade Republic returned a different page shape. Re-run with --login or --debug-dir tr_debug."
                )
                break

        context.close()

    if abort_due_to_empty_pages:
        return 2

    workbook.save(workbook_path)
    print(f"Saved: {workbook_path}")
    print(f"Updated rows: {len(data_rows)} | TRUE: {true_count} | FALSE: {false_count}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
