"""Add a research-informed Final_Score column to stocks_metrics.xlsx.

The score is a region-relative, cross-sectional ranking intended for long-term
screening. It is not a price target or a guarantee of future returns.
"""

from __future__ import annotations

import argparse
import os
import shutil
import tempfile
from copy import copy
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.table import TableColumn


DEFAULT_FILE = Path("stocks_metrics.xlsx")
DEFAULT_SHEET = "Metrics"
FINAL_SCORE_COLUMN = "Final_Score"
MIN_COVERAGE = 0.60
WINSOR_LOWER = 0.05
WINSOR_UPPER = 0.95

# Weights sum to 1.00. Sentiment_Score is deliberately excluded because the
# existing implementation already contains momentum and fundamental inputs.
FACTOR_WEIGHTS = {
    "profitability": 0.25,
    "value": 0.25,
    "growth_quality": 0.20,
    "momentum": 0.15,
    "balance_sheet": 0.10,
    "analyst": 0.05,
}

REQUIRED_COLUMNS = {
    "Region",
    "PE_Ratio",
    "Profit_Margin",
    "Revenue_Growth",
    "1Y_Momentum",
    "Debt_To_Equity",
    "Analyst_Rating",
}


def _finite_numeric(series: pd.Series) -> pd.Series:
    """Coerce a series to finite floating-point values."""
    numeric = pd.to_numeric(series, errors="coerce").astype(float)
    return numeric.where(np.isfinite(numeric))


def _sanitize_metrics(df: pd.DataFrame) -> pd.DataFrame:
    """Convert metrics to numeric values and reject structurally invalid data."""
    clean = pd.DataFrame(index=df.index)
    clean["Region"] = df["Region"].fillna("Unknown").astype(str).str.strip()

    clean["profitability"] = _finite_numeric(df["Profit_Margin"]).where(
        lambda s: s.between(-5.0, 5.0)
    )
    clean["revenue_growth"] = _finite_numeric(df["Revenue_Growth"]).where(
        lambda s: s.between(-1.0, 10.0)
    )
    clean["momentum"] = _finite_numeric(df["1Y_Momentum"]).where(
        lambda s: s.between(-100.0, 10_000.0)
    )
    clean["pe_ratio"] = _finite_numeric(df["PE_Ratio"]).where(
        lambda s: s.between(0.01, 500.0)
    )
    # Yahoo reports debt-to-equity as a percentage (e.g. 150 means 150%).
    # Negative values generally reflect negative equity, not conservative debt.
    clean["debt_to_equity"] = _finite_numeric(df["Debt_To_Equity"]).where(
        lambda s: s.between(0.0, 5_000.0)
    )
    # Yahoo recommendationMean: 1 = Strong Buy and 5 = Sell.
    clean["analyst_rating"] = _finite_numeric(df["Analyst_Rating"]).where(
        lambda s: s.between(1.0, 5.0)
    )
    return clean


def _percentile_score(
    values: pd.Series,
    regions: pd.Series,
    *,
    higher_is_better: bool,
) -> pd.Series:
    """Winsorize and percentile-rank values within each region."""
    result = pd.Series(np.nan, index=values.index, dtype=float)

    for region in regions.unique():
        region_index = regions.index[regions == region]
        region_values = values.loc[region_index].dropna()
        if region_values.empty:
            continue

        lower = region_values.quantile(WINSOR_LOWER)
        upper = region_values.quantile(WINSOR_UPPER)
        clipped = region_values.clip(lower=lower, upper=upper)

        ascending = higher_is_better
        ranks = clipped.rank(method="average", ascending=ascending)
        if len(clipped) == 1:
            scores = pd.Series(0.5, index=clipped.index)
        else:
            scores = (ranks - 1.0) / (len(clipped) - 1.0)
        result.loc[scores.index] = scores

    return result


def calculate_final_scores(df: pd.DataFrame) -> tuple[pd.Series, pd.DataFrame]:
    """Return Final_Score and an auditable factor-score DataFrame."""
    missing = sorted(REQUIRED_COLUMNS.difference(df.columns))
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")

    clean = _sanitize_metrics(df)
    regions = clean["Region"]

    factors = pd.DataFrame(index=df.index)
    factors["profitability"] = _percentile_score(
        clean["profitability"], regions, higher_is_better=True
    )
    factors["value"] = _percentile_score(
        clean["pe_ratio"], regions, higher_is_better=False
    )
    revenue_score = _percentile_score(
        clean["revenue_growth"], regions, higher_is_better=True
    )
    factors["growth_quality"] = revenue_score * (
        0.5 + 0.5 * factors["profitability"]
    )
    factors["momentum"] = _percentile_score(
        clean["momentum"], regions, higher_is_better=True
    )
    factors["balance_sheet"] = _percentile_score(
        clean["debt_to_equity"], regions, higher_is_better=False
    )
    factors["analyst"] = _percentile_score(
        clean["analyst_rating"], regions, higher_is_better=False
    )

    weight_series = pd.Series(FACTOR_WEIGHTS)
    available_weight = factors.notna().mul(weight_series, axis=1).sum(axis=1)
    weighted_points = factors.mul(weight_series, axis=1).sum(axis=1, skipna=True)

    base_score = weighted_points.div(available_weight.where(available_weight > 0))
    completeness_multiplier = 0.85 + 0.15 * available_weight
    final_score = 100.0 * base_score * completeness_multiplier
    final_score = final_score.where(available_weight >= MIN_COVERAGE).round(1)

    factors["coverage"] = available_weight
    factors["final_score"] = final_score
    return final_score, factors


def _copy_cell_style(source, destination) -> None:
    if source.has_style:
        destination._style = copy(source._style)
    if source.number_format:
        destination.number_format = source.number_format
    destination.font = copy(source.font)
    destination.fill = copy(source.fill)
    destination.border = copy(source.border)
    destination.alignment = copy(source.alignment)
    destination.protection = copy(source.protection)


def _extend_sheet_ranges(worksheet, score_column: int) -> None:
    """Extend tables and filters to include a newly appended score column."""
    for table in worksheet.tables.values():
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        if min_col <= score_column and max_col == score_column - 1:
            next_id = max((column.id for column in table.tableColumns), default=0) + 1
            table.tableColumns.append(
                TableColumn(id=next_id, name=FINAL_SCORE_COLUMN)
            )
            table.ref = (
                f"{get_column_letter(min_col)}{min_row}:"
                f"{get_column_letter(score_column)}{max_row}"
            )
            if table.autoFilter is not None:
                table.autoFilter.ref = table.ref

    if worksheet.auto_filter.ref:
        min_col, min_row, max_col, max_row = range_boundaries(
            worksheet.auto_filter.ref
        )
        if min_col <= score_column and max_col == score_column - 1:
            worksheet.auto_filter.ref = (
                f"{get_column_letter(min_col)}{min_row}:"
                f"{get_column_letter(score_column)}{max_row}"
            )


def write_scores(
    source_path: Path,
    output_path: Path,
    sheet_name: str,
    scores: pd.Series,
) -> None:
    """Write scores while preserving the rest of the workbook."""
    if source_path.resolve() != output_path.resolve():
        output_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(source_path, output_path)

    workbook = load_workbook(output_path)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Worksheet '{sheet_name}' not found in {output_path}")
    worksheet = workbook[sheet_name]

    headers = {
        str(cell.value).strip(): cell.column
        for cell in worksheet[1]
        if cell.value is not None
    }
    is_new_column = FINAL_SCORE_COLUMN not in headers
    score_column = headers.get(FINAL_SCORE_COLUMN, worksheet.max_column + 1)

    if is_new_column and score_column > 1:
        _copy_cell_style(
            worksheet.cell(row=1, column=score_column - 1),
            worksheet.cell(row=1, column=score_column),
        )
        _extend_sheet_ranges(worksheet, score_column)

    header = worksheet.cell(row=1, column=score_column)
    header.value = FINAL_SCORE_COLUMN
    header.comment = Comment(
        "Region-relative long-term screening score (0-100). "
        "Weights: profitability 25%, value 25%, growth quality 20%, "
        "1Y momentum 15%, balance sheet 10%, analyst rating 5%. "
        "Inputs are winsorized at the 5th/95th percentiles. Scores with "
        "less than 60% factor coverage are left blank.",
        "FinSentiment",
    )

    for offset, score in enumerate(scores, start=2):
        destination = worksheet.cell(row=offset, column=score_column)
        if is_new_column and score_column > 1:
            _copy_cell_style(
                worksheet.cell(row=offset, column=score_column - 1),
                destination,
            )
        destination.value = None if pd.isna(score) else float(score)
        destination.number_format = "0.0"

    score_letter = get_column_letter(score_column)
    worksheet.column_dimensions[score_letter].width = max(
        worksheet.column_dimensions[score_letter].width or 0,
        len(FINAL_SCORE_COLUMN) + 2,
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    temp_handle = tempfile.NamedTemporaryFile(
        prefix=f".{output_path.stem}-",
        suffix=".xlsx",
        dir=output_path.parent,
        delete=False,
    )
    temp_path = Path(temp_handle.name)
    temp_handle.close()
    try:
        workbook.save(temp_path)
        os.replace(temp_path, output_path)
    finally:
        if temp_path.exists():
            temp_path.unlink()


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Calculate and append Final_Score to a metrics workbook."
    )
    parser.add_argument("--input", type=Path, default=DEFAULT_FILE)
    parser.add_argument(
        "--output",
        type=Path,
        help="Optional output workbook. Defaults to updating --input in place.",
    )
    parser.add_argument("--sheet", default=DEFAULT_SHEET)
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Calculate and summarize scores without writing the workbook.",
    )
    return parser


def main() -> int:
    args = _build_parser().parse_args()
    source_path = args.input.resolve()
    output_path = (args.output or args.input).resolve()

    if not source_path.exists():
        raise FileNotFoundError(
            f"{source_path} does not exist. Run stock_sentiment.py first."
        )

    dataframe = pd.read_excel(source_path, sheet_name=args.sheet)
    scores, audit = calculate_final_scores(dataframe)

    scored = int(scores.notna().sum())
    unscored = int(scores.isna().sum())
    print(f"Rows: {len(dataframe)} | scored: {scored} | unscored: {unscored}")
    if scored:
        print(
            "Final_Score distribution: "
            f"min={scores.min():.1f}, median={scores.median():.1f}, "
            f"max={scores.max():.1f}"
        )
    print(
        "Average factor coverage: "
        f"{audit['coverage'].mean() * 100:.1f}%"
    )

    if args.dry_run:
        print("Dry run complete; workbook was not changed.")
        return 0

    write_scores(source_path, output_path, args.sheet, scores)
    print(f"Saved Final_Score to {output_path} (sheet: {args.sheet})")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"Error: {exc}")
        raise SystemExit(1)
